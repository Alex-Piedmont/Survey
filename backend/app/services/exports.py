"""Export service for CSV and XLSX generation."""

import csv
import io
from datetime import date

from openpyxl import Workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.models.presentation_grade import PresentationGrade
from app.models.session import Session, SessionTeam
from app.models.survey import Question, SurveyTemplate
from app.models.team import Team, TeamMembership
from app.services.aggregations import (
    get_enrolled_student_emails,
    get_latest_submissions,
    get_presenting_member_emails,
)


async def _get_session_context(db: AsyncSession, session_id: str) -> dict:
    """Gather all context needed for export."""
    result = await db.execute(
        select(Session)
        .options(selectinload(Session.session_teams))
        .where(Session.id == session_id)
    )
    session = result.scalar_one()

    # Team info
    team_ids = [st.team_id for st in session.session_teams]
    teams_result = await db.execute(
        select(Team).options(selectinload(Team.memberships)).where(Team.id.in_(team_ids))
    )
    teams = {t.id: t for t in teams_result.scalars().all()}
    team_names = {tid: t.name for tid, t in teams.items()}

    # Template questions
    q_result = await db.execute(
        select(Question)
        .where(Question.template_id == session.template_snapshot_id, Question.is_active == True)
        .order_by(Question.sort_order)
    )
    questions = list(q_result.scalars().all())

    # Enrolled students
    enrolled = await get_enrolled_student_emails(db, session.section_id)
    presenters = await get_presenting_member_emails(db, session_id)

    # All latest submissions
    all_subs = await get_latest_submissions(db, session_id)

    # Presentation grades
    grades_result = await db.execute(
        select(PresentationGrade).where(PresentationGrade.session_id == session_id)
    )
    grades = {g.team_id: g for g in grades_result.scalars().all()}

    return {
        "session": session,
        "team_ids": team_ids,
        "team_names": team_names,
        "teams": teams,
        "questions": questions,
        "enrolled": enrolled,
        "presenters": presenters,
        "all_subs": all_subs,
        "grades": grades,
    }


def _build_participation_sheet(ctx: dict) -> list[list]:
    """Sheet 1: Per-student participation and penalty summary."""
    team_ids = ctx["team_ids"]
    team_names = ctx["team_names"]

    # Group subs by student
    student_subs: dict[str, list] = {}
    for sub in ctx["all_subs"]:
        student_subs.setdefault(sub.student_email, []).append(sub)

    # Header
    header = ["Student Email", "Is Presenter"]
    for tid in team_ids:
        header.append(f"Audience: {team_names[tid]}")
    header.extend(["Peer Submitted", "Late Submissions", "Max Penalty %"])

    rows = [header]
    for email in sorted(ctx["enrolled"]):
        subs = student_subs.get(email, [])
        is_presenter = email in ctx["presenters"]

        # Audience completion per team
        audience_by_team = {}
        peer_done = False
        late_count = 0
        max_penalty = 0
        for s in subs:
            if s.feedback_type == "audience" and s.target_team_id:
                audience_by_team[s.target_team_id] = "on-time" if not s.is_late else f"late ({s.penalty_pct}%)"
            elif s.feedback_type == "peer":
                peer_done = True
            if s.is_late:
                late_count += 1
                max_penalty = max(max_penalty, s.penalty_pct)

        row = [email, "Yes" if is_presenter else "No"]
        for tid in team_ids:
            # Presenters can't rate own team
            own_team = any(
                m.student_email == email
                for m in ctx["teams"].get(tid, Team()).memberships
                if hasattr(ctx["teams"].get(tid), "memberships")
            )
            if own_team:
                row.append("N/A (own team)")
            else:
                row.append(audience_by_team.get(tid, "missing"))
        row.extend([
            "Yes" if peer_done else ("N/A" if not is_presenter else "missing"),
            str(late_count),
            str(max_penalty),
        ])
        rows.append(row)

    return rows


def _build_audience_sheet(ctx: dict) -> list[list]:
    """Sheet 2: Audience feedback raw responses."""
    audience_qs = [q for q in ctx["questions"] if q.category == "audience"]
    if not audience_qs:
        return [["No audience questions in template"]]

    header = ["Submitter", "Target Team"]
    for q in audience_qs:
        header.append(q.question_text)
    header.extend(["Submitted At", "Is Late", "Penalty %", "Withheld"])

    rows = [header]
    for sub in sorted(ctx["all_subs"], key=lambda s: (s.target_team_id or "", s.student_email)):
        if sub.feedback_type != "audience":
            continue
        row = [
            sub.student_email,
            ctx["team_names"].get(sub.target_team_id, "Unknown"),
        ]
        for q in audience_qs:
            row.append(sub.responses.get(q.id, ""))
        submitted_at = sub.submitted_at.isoformat() if sub.submitted_at else ""
        row.extend([submitted_at, "Yes" if sub.is_late else "No", str(sub.penalty_pct), "Yes" if sub.withheld else "No"])
        rows.append(row)

    return rows


def _build_peer_sheet(ctx: dict) -> list[list]:
    """Sheet 3: Peer feedback raw responses."""
    peer_qs = [q for q in ctx["questions"] if q.category == "peer"]
    if not peer_qs:
        return [["No peer questions in template"]]

    header = ["Submitter", "Target Student", "Team"]
    for q in peer_qs:
        header.append(q.question_text)
    header.extend(["Submitted At", "Is Late", "Penalty %"])

    rows = [header]
    for sub in sorted(ctx["all_subs"], key=lambda s: (s.target_team_id or "", s.student_email)):
        if sub.feedback_type != "peer":
            continue
        row = [
            sub.student_email,
            sub.target_student_email or "",
            ctx["team_names"].get(sub.target_team_id, "Unknown"),
        ]
        for q in peer_qs:
            row.append(sub.responses.get(q.id, ""))
        submitted_at = sub.submitted_at.isoformat() if sub.submitted_at else ""
        row.extend([submitted_at, "Yes" if sub.is_late else "No", str(sub.penalty_pct)])
        rows.append(row)

    return rows


def _build_instructor_sheet(ctx: dict) -> list[list]:
    """Sheet 4: Instructor feedback and presentation quality grades."""
    rows = [["Team", "Presentation Grade", "Grade Comments"]]

    # Instructor submissions
    instructor_subs = [s for s in ctx["all_subs"] if s.feedback_type == "instructor"]

    for tid in ctx["team_ids"]:
        grade = ctx["grades"].get(tid)
        rows.append([
            ctx["team_names"].get(tid, "Unknown"),
            grade.grade if grade else "",
            grade.comments or "" if grade else "",
        ])

    if instructor_subs:
        rows.append([])
        rows.append(["--- Instructor Feedback Responses ---"])
        header = ["Target Team"]
        # Gather all response keys
        all_keys = set()
        for sub in instructor_subs:
            all_keys.update(sub.responses.keys())
        sorted_keys = sorted(all_keys)
        header.extend(sorted_keys)
        rows.append(header)

        for sub in instructor_subs:
            row = [ctx["team_names"].get(sub.target_team_id, "Unknown")]
            for k in sorted_keys:
                row.append(sub.responses.get(k, ""))
            rows.append(row)

    return rows


async def export_session_csv(db: AsyncSession, session_id: str) -> str:
    """Generate CSV export for a session. Returns CSV string."""
    ctx = await _get_session_context(db, session_id)

    output = io.StringIO()
    writer = csv.writer(output)

    writer.writerow(["=== Participation Summary ==="])
    for row in _build_participation_sheet(ctx):
        writer.writerow(row)

    writer.writerow([])
    writer.writerow(["=== Audience Feedback ==="])
    for row in _build_audience_sheet(ctx):
        writer.writerow(row)

    writer.writerow([])
    writer.writerow(["=== Peer Feedback ==="])
    for row in _build_peer_sheet(ctx):
        writer.writerow(row)

    writer.writerow([])
    writer.writerow(["=== Instructor Feedback & Grades ==="])
    for row in _build_instructor_sheet(ctx):
        writer.writerow(row)

    return output.getvalue()


async def export_session_xlsx(db: AsyncSession, session_id: str) -> bytes:
    """Generate XLSX export for a session. Returns bytes."""
    ctx = await _get_session_context(db, session_id)

    wb = Workbook()

    # Sheet 1: Participation
    ws1 = wb.active
    ws1.title = "Participation Summary"
    for row in _build_participation_sheet(ctx):
        ws1.append(row)

    # Sheet 2: Audience
    ws2 = wb.create_sheet("Audience Feedback")
    for row in _build_audience_sheet(ctx):
        ws2.append(row)

    # Sheet 3: Peer
    ws3 = wb.create_sheet("Peer Feedback")
    for row in _build_peer_sheet(ctx):
        ws3.append(row)

    # Sheet 4: Instructor
    ws4 = wb.create_sheet("Instructor & Grades")
    for row in _build_instructor_sheet(ctx):
        ws4.append(row)

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()
