import csv
import io

import pytest
from httpx import AsyncClient
from openpyxl import load_workbook

from app.core.security import create_access_token


async def _setup_session_with_submissions(client: AsyncClient, auth_headers: dict) -> dict:
    """Create full setup with some submissions for export testing."""
    # Course
    resp = await client.post(
        "/api/v1/courses",
        json={"name": "MGMT 481", "term": "Spring 2026"},
        headers=auth_headers,
    )
    course_id = resp.json()["id"]

    # Section
    resp = await client.post(
        f"/api/v1/courses/{course_id}/sections",
        json={"name": "Section A"},
        headers=auth_headers,
    )
    section_id = resp.json()["id"]

    # Enroll students
    await client.post(
        f"/api/v1/sections/{section_id}/enroll",
        json={"emails": "alice@uni.edu\nbob@uni.edu\neve@uni.edu"},
        headers=auth_headers,
    )

    # Presentation type + template
    resp = await client.post(
        f"/api/v1/courses/{course_id}/presentation-types",
        json={"name": "Debates"},
        headers=auth_headers,
    )
    ptype_id = resp.json()["id"]

    await client.put(
        f"/api/v1/presentation-types/{ptype_id}/template",
        json={
            "questions": [
                {
                    "question_text": "Content clarity",
                    "question_type": "likert_5",
                    "category": "audience",
                    "sort_order": 1,
                },
                {
                    "question_text": "Comments",
                    "question_type": "free_text",
                    "category": "audience",
                    "is_required": False,
                    "sort_order": 2,
                },
                {
                    "question_text": "Contribution",
                    "question_type": "likert_5",
                    "category": "peer",
                    "sort_order": 3,
                },
            ]
        },
        headers=auth_headers,
    )

    # Team: alice, bob
    resp = await client.post(
        f"/api/v1/sections/{section_id}/teams",
        json={
            "name": "Team Alpha",
            "presentation_type_id": ptype_id,
            "member_emails": ["alice@uni.edu", "bob@uni.edu"],
        },
        headers=auth_headers,
    )
    team_id = resp.json()["id"]

    # Session
    resp = await client.post(
        "/api/v1/sessions",
        json={
            "section_id": section_id,
            "presentation_type_id": ptype_id,
            "presenting_team_ids": [team_id],
            "session_date": "2026-12-15",
        },
        headers=auth_headers,
    )
    session_id = resp.json()["id"]

    # Get question IDs
    resp = await client.get(f"/api/v1/s/{session_id}")
    questions = resp.json()["questions"]
    likert_qid = next(q["id"] for q in questions if q["question_type"].startswith("likert") and q["category"] == "audience")
    text_qid = next(q["id"] for q in questions if q["question_type"] == "free_text")
    peer_qid = next(q["id"] for q in questions if q["category"] == "peer")

    # Eve submits audience feedback
    eve_headers = {"Authorization": f"Bearer {create_access_token('eve@uni.edu')}"}
    await client.post(
        f"/api/v1/s/{session_id}/submit",
        json={
            "target_team_id": team_id,
            "feedback_type": "audience",
            "responses": {likert_qid: 4, text_qid: "Good work!"},
        },
        headers=eve_headers,
    )

    # Alice submits peer feedback for Bob
    alice_headers = {"Authorization": f"Bearer {create_access_token('alice@uni.edu')}"}
    await client.post(
        f"/api/v1/s/{session_id}/submit",
        json={
            "target_student_email": "bob@uni.edu",
            "feedback_type": "peer",
            "responses": {peer_qid: 5},
        },
        headers=alice_headers,
    )

    # Instructor assigns a grade
    await client.post(
        f"/api/v1/sessions/{session_id}/teams/{team_id}/presentation-grade",
        json={"grade": "A", "comments": "Great presentation"},
        headers=auth_headers,
    )

    return {
        "session_id": session_id,
        "section_id": section_id,
        "team_id": team_id,
    }


class TestCSVExport:
    async def test_csv_export(self, client: AsyncClient, auth_headers: dict):
        ids = await _setup_session_with_submissions(client, auth_headers)

        resp = await client.get(
            f"/api/v1/sessions/{ids['session_id']}/export?format=csv",
            headers=auth_headers,
        )
        assert resp.status_code == 200
        assert "text/csv" in resp.headers["content-type"]

        content = resp.text
        reader = csv.reader(io.StringIO(content))
        rows = list(reader)

        # Should have participation, audience, peer, and instructor sections
        section_headers = [r[0] for r in rows if r and r[0].startswith("===")]
        assert "=== Participation Summary ===" in section_headers
        assert "=== Audience Feedback ===" in section_headers
        assert "=== Peer Feedback ===" in section_headers
        assert "=== Instructor Feedback & Grades ===" in section_headers

    async def test_csv_contains_student_data(self, client: AsyncClient, auth_headers: dict):
        ids = await _setup_session_with_submissions(client, auth_headers)

        resp = await client.get(
            f"/api/v1/sessions/{ids['session_id']}/export?format=csv",
            headers=auth_headers,
        )
        content = resp.text

        # Should contain student emails
        assert "eve@uni.edu" in content
        assert "alice@uni.edu" in content
        assert "bob@uni.edu" in content

    async def test_csv_requires_instructor(self, client: AsyncClient, auth_headers: dict):
        ids = await _setup_session_with_submissions(client, auth_headers)
        student_headers = {"Authorization": f"Bearer {create_access_token('eve@uni.edu')}"}

        resp = await client.get(
            f"/api/v1/sessions/{ids['session_id']}/export?format=csv",
            headers=student_headers,
        )
        assert resp.status_code == 403


class TestXLSXExport:
    async def test_xlsx_export(self, client: AsyncClient, auth_headers: dict):
        ids = await _setup_session_with_submissions(client, auth_headers)

        resp = await client.get(
            f"/api/v1/sessions/{ids['session_id']}/export?format=xlsx",
            headers=auth_headers,
        )
        assert resp.status_code == 200
        assert "spreadsheetml" in resp.headers["content-type"]

        # Parse the XLSX
        wb = load_workbook(io.BytesIO(resp.content))
        assert "Participation Summary" in wb.sheetnames
        assert "Audience Feedback" in wb.sheetnames
        assert "Peer Feedback" in wb.sheetnames
        assert "Instructor & Grades" in wb.sheetnames

    async def test_xlsx_participation_sheet(self, client: AsyncClient, auth_headers: dict):
        ids = await _setup_session_with_submissions(client, auth_headers)

        resp = await client.get(
            f"/api/v1/sessions/{ids['session_id']}/export?format=xlsx",
            headers=auth_headers,
        )
        wb = load_workbook(io.BytesIO(resp.content))
        ws = wb["Participation Summary"]

        # Header row
        headers = [cell.value for cell in ws[1]]
        assert "Student Email" in headers
        assert "Is Presenter" in headers

        # Should have 3 students + header
        assert ws.max_row >= 4

    async def test_xlsx_audience_sheet(self, client: AsyncClient, auth_headers: dict):
        ids = await _setup_session_with_submissions(client, auth_headers)

        resp = await client.get(
            f"/api/v1/sessions/{ids['session_id']}/export?format=xlsx",
            headers=auth_headers,
        )
        wb = load_workbook(io.BytesIO(resp.content))
        ws = wb["Audience Feedback"]

        # Header + 1 submission from eve
        assert ws.max_row >= 2
        headers = [cell.value for cell in ws[1]]
        assert "Submitter" in headers
        assert "Content clarity" in headers

    async def test_xlsx_instructor_sheet_has_grade(self, client: AsyncClient, auth_headers: dict):
        ids = await _setup_session_with_submissions(client, auth_headers)

        resp = await client.get(
            f"/api/v1/sessions/{ids['session_id']}/export?format=xlsx",
            headers=auth_headers,
        )
        wb = load_workbook(io.BytesIO(resp.content))
        ws = wb["Instructor & Grades"]

        # Should contain the grade "A"
        all_values = []
        for row in ws.iter_rows(values_only=True):
            all_values.extend(row)
        assert "A" in all_values


class TestExportEdgeCases:
    async def test_export_empty_session(self, client: AsyncClient, auth_headers: dict):
        """Export works even with no submissions."""
        # Minimal setup without submissions
        resp = await client.post(
            "/api/v1/courses",
            json={"name": "Empty Course", "term": "Spring 2026"},
            headers=auth_headers,
        )
        course_id = resp.json()["id"]

        resp = await client.post(
            f"/api/v1/courses/{course_id}/sections",
            json={"name": "Section B"},
            headers=auth_headers,
        )
        section_id = resp.json()["id"]

        resp = await client.post(
            f"/api/v1/courses/{course_id}/presentation-types",
            json={"name": "Talks"},
            headers=auth_headers,
        )
        ptype_id = resp.json()["id"]

        await client.put(
            f"/api/v1/presentation-types/{ptype_id}/template",
            json={
                "questions": [
                    {"question_text": "Score", "question_type": "likert_5", "category": "audience", "sort_order": 1},
                ]
            },
            headers=auth_headers,
        )

        await client.post(
            f"/api/v1/sections/{section_id}/enroll",
            json={"emails": "test@uni.edu"},
            headers=auth_headers,
        )

        resp = await client.post(
            f"/api/v1/sections/{section_id}/teams",
            json={"name": "Solo", "presentation_type_id": ptype_id, "member_emails": ["test@uni.edu"]},
            headers=auth_headers,
        )
        team_id = resp.json()["id"]

        resp = await client.post(
            "/api/v1/sessions",
            json={
                "section_id": section_id,
                "presentation_type_id": ptype_id,
                "presenting_team_ids": [team_id],
                "session_date": "2026-12-15",
            },
            headers=auth_headers,
        )
        session_id = resp.json()["id"]

        # CSV export should work
        resp = await client.get(
            f"/api/v1/sessions/{session_id}/export?format=csv",
            headers=auth_headers,
        )
        assert resp.status_code == 200

        # XLSX export should work
        resp = await client.get(
            f"/api/v1/sessions/{session_id}/export?format=xlsx",
            headers=auth_headers,
        )
        assert resp.status_code == 200
